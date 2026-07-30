import os
import re
import sys
import pandas as pd
import pdfplumber
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DELIMITADOR = "|||"


# ==============================================================================
# FUNCIONES AUXILIARES COMUNES
# ==============================================================================
def obtener_ruta_asset(ruta_relativa):
    """Obtiene la ruta absoluta para assets, funciona en desarrollo y en el .exe de PyInstaller."""
    if hasattr(sys, "_MEIPASS"):
        return os.path.join(sys._MEIPASS, ruta_relativa)
    return os.path.join(os.path.abspath("."), ruta_relativa)


def es_numero_financiero(texto):
    """Valida si un texto es un monto monetario (positivo, negativo o decimales)."""
    if not texto:
        return False
    txt = texto.strip().replace("$", "").replace(" ", "")
    patron = r"^-?[\d\.,]+$"
    return bool(re.match(patron, txt))


_PATRON_PAG_TOKEN = re.compile(r"(?i)^p[áa]g(?:ina)?\.?$")
_PATRON_DE_TOKEN = re.compile(r"(?i)^de\.?$")
_PATRON_NUM_TOKEN = re.compile(r"^\d+$")


def limpiar_pie_pagina(words_fila):
    """Elimina la secuencia 'Página N de M' (o variantes) sin importar si comparte fila con datos reales."""
    n = len(words_fila)
    idx_a_quitar = set()
    i = 0
    while i < n:
        if _PATRON_PAG_TOKEN.match(words_fila[i]["text"].strip()):
            seq = [i]
            j = i + 1
            if j < n and _PATRON_NUM_TOKEN.match(words_fila[j]["text"].strip()):
                seq.append(j)
                j += 1
                if j < n and _PATRON_DE_TOKEN.match(
                    words_fila[j]["text"].strip()
                ):
                    seq.append(j)
                    j += 1
                    if j < n and _PATRON_NUM_TOKEN.match(
                        words_fila[j]["text"].strip()
                    ):
                        seq.append(j)
                        j += 1
            idx_a_quitar.update(seq)
            i = j
        else:
            i += 1
    if not idx_a_quitar:
        return words_fila
    return [w for k, w in enumerate(words_fila) if k not in idx_a_quitar]


def texto_delimitado_a_excel(lineas_texto, columnas, output_excel_path):
    """Convierte las líneas delimitadas a un DataFrame y lo exporta a Excel."""
    registros = []
    for linea in lineas_texto:
        partes = linea.split(DELIMITADOR)
        if len(partes) == len(columnas):
            registros.append(dict(zip(columnas, partes)))

    df = pd.DataFrame(registros, columns=columnas)
    df.to_excel(output_excel_path, index=False)


# ==============================================================================
# PARSER 1: MOVIMIENTOS SOCIEDAD
# ==============================================================================


def extraer_lineas_movimientos_soc(pdf_path):
    """Extrae transacciones para Movimientos Sociedad (7 columnas)."""
    patron_fecha = r"^\b(?:\d{4}/\d{2}/\d{2}|\d{1,2}/\d{2}(?:/\d{2,4})?)\b"
    lineas_delimitadas = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            words = page.extract_words()
            if not words:
                continue

            anc_pagina = page.width
            x_max_desc = anc_pagina * 0.35
            x_max_sucursal = anc_pagina * 0.50
            x_max_ref1 = anc_pagina * 0.63
            x_max_ref2 = anc_pagina * 0.73
            x_max_doc = anc_pagina * 0.85

            filas_y = {}
            for w in words:
                y_key = round(w["top"] / 3.0) * 3.0
                filas_y.setdefault(y_key, []).append(w)

            transacciones_pagina = []
            tx_actual = None

            for y_key in sorted(filas_y.keys()):
                words_brutas = sorted(filas_y[y_key], key=lambda x: x["x0"])
                words_linea = limpiar_pie_pagina(words_brutas)

                if not words_linea:
                    continue

                primera_word = words_linea[0]
                texto_primera = primera_word["text"].strip()

                es_inicio_tx = bool(
                    re.match(patron_fecha, texto_primera)
                    and primera_word["x0"] < (anc_pagina * 0.20)
                )

                if es_inicio_tx:
                    if tx_actual:
                        transacciones_pagina.append(tx_actual)

                    tx_actual = {
                        "FECHA": texto_primera,
                        "DESCRIPCIÓN": [],
                        "SUCURSAL/CANAL": [],
                        "REFERENCIA 1": [],
                        "REFERENCIA 2": [],
                        "DOCUMENTO": [],
                        "VALOR": "",
                    }
                    words_a_procesar = words_linea[1:]
                else:
                    if not tx_actual:
                        continue
                    words_a_procesar = words_linea

                for w in words_a_procesar:
                    x_centro = (w["x0"] + w["x1"]) / 2.0
                    txt_w = w["text"]

                    if x_centro < x_max_desc:
                        tx_actual["DESCRIPCIÓN"].append(txt_w)
                    elif x_centro < x_max_sucursal:
                        tx_actual["SUCURSAL/CANAL"].append(txt_w)
                    elif x_centro < x_max_ref1:
                        tx_actual["REFERENCIA 1"].append(txt_w)
                    elif x_centro < x_max_ref2:
                        tx_actual["REFERENCIA 2"].append(txt_w)
                    elif x_centro < x_max_doc:
                        tx_actual["DOCUMENTO"].append(txt_w)
                    else:
                        if es_numero_financiero(txt_w):
                            tx_actual["VALOR"] = txt_w

            if tx_actual:
                transacciones_pagina.append(tx_actual)

            for tx in transacciones_pagina:
                desc_str = " ".join(tx["DESCRIPCIÓN"]).strip()
                suc_str = " ".join(tx["SUCURSAL/CANAL"]).strip()
                ref1_str = " ".join(tx["REFERENCIA 1"]).strip()
                ref2_str = " ".join(tx["REFERENCIA 2"]).strip()
                doc_str = " ".join(tx["DOCUMENTO"]).strip()

                if (
                    "DESCRIPC" in desc_str.upper()
                    or "REFERENCIA" in ref1_str.upper()
                ):
                    continue

                registro = [
                    tx["FECHA"],
                    desc_str,
                    suc_str,
                    ref1_str,
                    ref2_str,
                    doc_str,
                    tx["VALOR"],
                ]
                lineas_delimitadas.append(DELIMITADOR.join(registro))

    return lineas_delimitadas


# ==============================================================================
# PARSER 2: MOVIMIENTOS PERSONA NATURAL
# ==============================================================================

def extraer_lineas_movimientos_pn(pdf_path):
    """Extrae transacciones para Movimientos Persona Natural de Bancolombia.

    Corrige la captura del año completo (evitando '202') y descarta metadatos de
    pie de página como 'Dirección IP'.
    """
    patron_fecha_pn = r"^\d{1,2}\s+(?:ene|feb|mar|abr|may|jun|jul|ago|sep|oct|nov|dic)\.?(?:\s+\d{4})?"
    lineas_delimitadas = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            words = page.extract_words()
            if not words:
                continue

            anc_pagina = page.width

            x_limite_desc = anc_pagina * 0.50
            x_limite_ref = anc_pagina * 0.76

            filas_y = {}
            for w in words:
                y_key = round(w["top"] / 3.0) * 3.0
                filas_y.setdefault(y_key, []).append(w)

            transacciones_pagina = []
            tx_actual = None

            for y_key in sorted(filas_y.keys()):
                words_brutas = sorted(filas_y[y_key], key=lambda x: x["x0"])
                words_linea = limpiar_pie_pagina(words_brutas)

                if not words_linea:
                    continue

                # --- 1. FILTRADO DE 'DIRECCIÓN IP' ---
                # Si la línea contiene metadatos de pie de página, los removemos antes de clasificar
                words_filtradas = []
                ignorar_siguientes = False
                for idx, w in enumerate(words_linea):
                    txt_clean = w["text"].strip().lower()
                    
                    # Detectar si inicia "Dirección IP"
                    if "direcci" in txt_clean or txt_clean == "ip" or "181.137" in txt_clean:
                        continue
                    words_filtradas.append(w)

                words_linea = words_filtradas
                if not words_linea:
                    continue

                # --- 2. CAPTURA Y VALIDACIÓN DE FECHA ---
                texto_linea_inicio = " ".join([w["text"] for w in words_linea[:4]])
                coincidencia_fecha = re.match(
                    patron_fecha_pn, texto_linea_inicio, re.IGNORECASE
                )

                es_inicio_tx = bool(
                    coincidencia_fecha
                    and words_linea[0]["x0"] < (anc_pagina * 0.25)
                )

                if es_inicio_tx:
                    if tx_actual:
                        transacciones_pagina.append(tx_actual)

                    fecha_str = coincidencia_fecha.group(0).strip()
                    num_words_fecha = len(fecha_str.split())

                    # Asegurar año de 4 dígitos si está presente inmediatamente después
                    if num_words_fecha < len(words_linea):
                        posible_anio = words_linea[num_words_fecha]["text"].strip()
                        if re.match(r"^\d{4}$", posible_anio):
                            fecha_str = f"{' '.join(fecha_str.split()[:2])} {posible_anio}"
                            num_words_fecha += 1

                    tx_actual = {
                        "FECHA": fecha_str,
                        "DESCRIPCIÓN": [],
                        "REFERENCIA": [],
                        "VALOR_RAW": [],
                    }
                    words_a_procesar = words_linea[num_words_fecha:]
                else:
                    if not tx_actual:
                        continue
                    words_a_procesar = words_linea

                # --- 3. CLASIFICACIÓN DE COLUMNAS POR COORDENADAS ---
                for w in words_a_procesar:
                    x_centro = (w["x0"] + w["x1"]) / 2.0
                    txt_w = w["text"].strip()

                    if not txt_w:
                        continue

                    if x_centro < x_limite_desc:
                        tx_actual["DESCRIPCIÓN"].append(txt_w)
                    elif x_centro < x_limite_ref:
                        tx_actual["REFERENCIA"].append(txt_w)
                    else:
                        tx_actual["VALOR_RAW"].append(txt_w)

            if tx_actual:
                transacciones_pagina.append(tx_actual)

            # --- 4. FORMATO FINAL Y LIMPIEZA DE VALORES ---
            for tx in transacciones_pagina:
                desc_str = " ".join(tx["DESCRIPCIÓN"]).strip()
                ref_str = " ".join(tx["REFERENCIA"]).strip()
                cadena_valor_bruta = "".join(tx["VALOR_RAW"]).strip()

                if (
                    "FECHA" in tx["FECHA"].upper()
                    or "DESCRIPCI" in desc_str.upper()
                ):
                    continue

                es_negativo = "-" in cadena_valor_bruta
                solo_num_y_puntos = re.sub(r"[^\d\.,]", "", cadena_valor_bruta)

                if "," in solo_num_y_puntos and "." in solo_num_y_puntos:
                    solo_num_y_puntos = solo_num_y_puntos.replace(".", "").replace(",", ".")
                elif "," in solo_num_y_puntos:
                    solo_num_y_puntos = solo_num_y_puntos.replace(",", ".")

                match_monto = re.search(r"\d+(?:\.\d+)?", solo_num_y_puntos)

                if match_monto:
                    monto_final = match_monto.group(0)
                    if es_negativo:
                        monto_final = f"-{monto_final}"
                else:
                    monto_final = "0"

                registro = [tx["FECHA"], desc_str, ref_str, monto_final]
                lineas_delimitadas.append(DELIMITADOR.join(registro))

    return lineas_delimitadas

# ==============================================================================
# PARSER 3: EXTRACTOS
# ==============================================================================


def extraer_lineas_extractos(pdf_path):
    """Extrae las transacciones para archivos de tipo 'Extracto'."""
    patron_fecha = r"^\b\d{1,2}/\d{2}(?:/\d{2,4})?\b"
    lineas_delimitadas = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            words = page.extract_words()
            if not words:
                continue

            anc_pagina = page.width
            x_max_desc = anc_pagina * 0.40
            x_max_sucursal = anc_pagina * 0.65

            filas_y = {}
            for w in words:
                y_key = round(w["top"] / 2.5) * 2.5
                filas_y.setdefault(y_key, []).append(w)

            for y_key in sorted(filas_y.keys()):
                words_ordenadas = sorted(filas_y[y_key], key=lambda x: x["x0"])
                words_linea = limpiar_pie_pagina(words_ordenadas)

                if not words_linea:
                    continue

                primera_word = words_linea[0]
                texto_primera = primera_word["text"].strip()

                if re.match(patron_fecha, texto_primera) and primera_word[
                    "x0"
                ] < (anc_pagina * 0.20):
                    fecha_val = texto_primera
                    resto_words = words_linea[1:]
                    if not resto_words:
                        continue

                    saldo_val = ""
                    valor_val = ""

                    if len(resto_words) > 0 and es_numero_financiero(
                        resto_words[-1]["text"]
                    ):
                        saldo_val = resto_words.pop(-1)["text"]

                    if len(resto_words) > 0 and es_numero_financiero(
                        resto_words[-1]["text"]
                    ):
                        valor_val = resto_words.pop(-1)["text"]

                    descripcion_words = []
                    sucursal_words = []
                    dcto_words = []

                    for w in resto_words:
                        x_centro = (w["x0"] + w["x1"]) / 2.0
                        txt_w = w["text"]

                        if x_centro < x_max_desc:
                            descripcion_words.append(txt_w)
                        elif x_centro < x_max_sucursal:
                            sucursal_words.append(txt_w)
                        else:
                            dcto_words.append(txt_w)

                    descripcion_val = " ".join(descripcion_words).strip()
                    sucursal_val = " ".join(sucursal_words).strip()
                    dcto_val = " ".join(dcto_words).strip()

                    if (
                        "DESCRIPC" in descripcion_val.upper()
                        or "RESUMEN" in descripcion_val.upper()
                    ):
                        continue

                    registro = [
                        fecha_val,
                        descripcion_val,
                        sucursal_val,
                        dcto_val,
                        valor_val,
                        saldo_val,
                    ]
                    lineas_delimitadas.append(DELIMITADOR.join(registro))

    return lineas_delimitadas


# ==============================================================================
# REORGANIZACIÓN Y GENERACIÓN DE REPORTES EN EXCEL
# ==============================================================================


def reorganizar_excel(excel_path):
    """Calcula totales, conciliación y conceptos. La hoja 'Resumen' omite saldos si no aplican."""
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"No se encontró el archivo: {excel_path}")

    xl = pd.ExcelFile(excel_path)
    sheet_a_leer = "Datos" if "Datos" in xl.sheet_names else 0
    df = pd.read_excel(excel_path, sheet_name=sheet_a_leer, dtype=str)

    if "VALOR" not in df.columns or df.empty:
        return excel_path

    col_desc = "DESCRIPCIÓN" if "DESCRIPCIÓN" in df.columns else "DESCRIPCION"

    # PASO 1: LIMPIEZA DE NÚMEROS
    valores_numericos = pd.to_numeric(
        df["VALOR"]
        .astype(str)
        .str.replace(",", "", regex=False)
        .str.replace("$", "", regex=False)
        .str.replace(" ", "", regex=False)
        .str.strip(),
        errors="coerce",
    ).fillna(0.0)

    tiene_columna_saldo = "SALDO" in df.columns
    saldos_numericos = pd.Series(dtype=float)

    if tiene_columna_saldo:
        saldos_numericos = pd.to_numeric(
            df["SALDO"]
            .astype(str)
            .str.replace(",", "", regex=False)
            .str.replace("$", "", regex=False)
            .str.replace(" ", "", regex=False)
            .str.strip(),
            errors="coerce",
        ).fillna(0.0)

    # PASO 2: CÁLCULOS SOBRE EL DF ORIGINAL
    cargos_sum = valores_numericos[valores_numericos < 0].sum()
    abonos_sum = valores_numericos[valores_numericos > 0].sum()

    filas_resumen = []
    if tiene_columna_saldo and not saldos_numericos.empty:
        # Corrección: El Saldo Anterior es el Saldo Inicial menos el Valor Ajustado del primer registro
        saldo_anterior = saldos_numericos.iloc[0] - valores_numericos.iloc[0]
        saldo_actual = saldos_numericos.iloc[-1]
        filas_resumen.append(
            {"CONCEPTO": "SALDO ANTERIOR", "MONTO": saldo_anterior}
        )

    filas_resumen.append({"CONCEPTO": "TOTAL CARGOS", "MONTO": cargos_sum})
    filas_resumen.append({"CONCEPTO": "TOTAL ABONOS", "MONTO": abonos_sum})

    if tiene_columna_saldo and not saldos_numericos.empty:
        filas_resumen.append(
            {"CONCEPTO": "SALDO ACTUAL", "MONTO": saldo_actual}
        )

    df_conciliacion = pd.DataFrame(filas_resumen)

    # Agrupar por Conceptos
    df_calc = df.copy()
    df_calc["_VALOR_NUM"] = valores_numericos
    df_calc["_CARGOS_TEMP"] = df_calc["_VALOR_NUM"].apply(
        lambda x: x if x < 0 else 0.0
    )
    df_calc["_ABONOS_TEMP"] = df_calc["_VALOR_NUM"].apply(
        lambda x: x if x > 0 else 0.0
    )

    df_conceptos = (
        df_calc.groupby(col_desc, as_index=False)
        .agg(
            CARGOS=("_CARGOS_TEMP", "sum"),
            ABONOS=("_ABONOS_TEMP", "sum"),
            NETO=("_VALOR_NUM", "sum"),
        )
        .sort_values(by="CARGOS", ascending=True)
    )

    fila_total = pd.DataFrame(
        [
            {
                col_desc: "TOTAL GENERAL",
                "CARGOS": float(df_conceptos["CARGOS"].sum()),
                "ABONOS": float(df_conceptos["ABONOS"].sum()),
                "NETO": float(df_conceptos["NETO"].sum()),
            }
        ]
    )

    df_conceptos = pd.concat([df_conceptos, fila_total], ignore_index=True)

    # PASO 3: REORGANIZACIÓN VISUAL (Negativos arriba)
    df_datos = df.copy()
    df_datos["VALOR"] = valores_numericos
    if tiene_columna_saldo:
        df_datos["SALDO"] = saldos_numericos

    is_negative = df_datos["VALOR"] < 0
    df_datos["_orden_temp"] = 0
    df_datos.loc[~is_negative, "_orden_temp"] = 1

    df_datos_ordenado = df_datos.sort_values(
        by="_orden_temp", kind="stable"
    ).drop(columns=["_orden_temp"])

    # PASO 4: EXPORTACIÓN Y FORMATO EN EXCEL
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df_datos_ordenado.to_excel(writer, sheet_name="Datos", index=False)
        df_conciliacion.to_excel(writer, sheet_name="Resumen", index=False)
        df_conceptos.to_excel(writer, sheet_name="Conceptos", index=False)

        workbook = writer.book
        FORMATO_MONEDA = '"$"#,##0.00;[Red]("$"#,##0.00);"-"'

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            for col in worksheet.iter_cols(1, worksheet.max_column):
                header_val = str(col[0].value).upper() if col[0].value else ""

                if header_val in [
                    "VALOR",
                    "SALDO",
                    "CARGOS",
                    "ABONOS",
                    "NETO",
                    "MONTO",
                ]:
                    for cell in col[1:]:
                        if cell.value is not None and isinstance(
                            cell.value, (int, float)
                        ):
                            cell.number_format = FORMATO_MONEDA

    return excel_path


# ==============================================================================
# ORQUESTADOR PRINCIPAL
# ==============================================================================


def ejecutar_proceso_exportacion(pdf_path, output_excel_path=None):
    """Detecta automáticamente el tipo de documento según el nombre del PDF."""
    nombre_archivo = os.path.basename(pdf_path).upper()

    if "MOVIMENTOSPNATURAL" in nombre_archivo or "MOVIMIENTOSPNATURAL" in nombre_archivo:
        tipo = "MOVIMIENTOS_PNATURAL"
        columnas = ["FECHA", "DESCRIPCIÓN", "REFERENCIA", "VALOR"]
        lineas_plana = extraer_lineas_movimientos_pn(pdf_path)

    elif "MOVIMIENTOSOC" in nombre_archivo or "MOVIMIENTO" in nombre_archivo:
        tipo = "MOVIMIENTOS_SOC"
        columnas = [
            "FECHA",
            "DESCRIPCIÓN",
            "SUCURSAL/CANAL",
            "REFERENCIA 1",
            "REFERENCIA 2",
            "DOCUMENTO",
            "VALOR",
        ]
        lineas_plana = extraer_lineas_movimientos_soc(pdf_path)

    elif "EXTRACTO" in nombre_archivo:
        tipo = "EXTRACTOS"
        columnas = [
            "FECHA",
            "DESCRIPCIÓN",
            "SUCURSAL",
            "DCTO.",
            "VALOR",
            "SALDO",
        ]
        lineas_plana = extraer_lineas_extractos(pdf_path)

    else:
        raise ValueError(
            "El nombre del archivo no coincide con un prefijo válido "
            "('MovimientosPNatural', 'MovimientoSOC' o 'Extracto')."
        )

    if not lineas_plana:
        return None

    if not output_excel_path:
        base_path, _ = os.path.splitext(pdf_path)
        output_excel_path = f"{base_path}_convertido.xlsx"

    base_txt_path, _ = os.path.splitext(output_excel_path)
    txt_path = f"{base_txt_path}_plano.txt"

    # 1. Guardar archivo Plano (.txt)
    encabezado_txt = DELIMITADOR.join(columnas)
    contenido_txt = [encabezado_txt] + lineas_plana

    with open(txt_path, "w", encoding="utf-8") as f:
        f.write("\n".join(contenido_txt))

    # 2. Convertir a Excel
    texto_delimitado_a_excel(lineas_plana, columnas, output_excel_path)

    # 3. Reorganizar y crear pestañas suplementarias
    reorganizar_excel(output_excel_path)

    return output_excel_path