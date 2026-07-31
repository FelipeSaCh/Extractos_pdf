import tkinter as tk
from tkinter import messagebox, ttk
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd


def aplicar_estilos_segun_imagen(excel_path):
    """Aplica la estética exacta de la imagen:

    - Punto (.) para miles, Coma (,) para decimales.
    - Negativos en ROJO entre paréntesis ($1.000,00).
    - Ceros presentados como un guion '-'.
    - Calibri 11, bordes delgados y encabezados limpios.
    """
    wb = openpyxl.load_workbook(excel_path)

    FORMATO_MONEDA_IMAGEN = (
        '"$"#,##0.00;[Red]("$"#,##0.00);"-"'
    )

    font_header = Font(name="Calibri", size=11, bold=True)
    font_body = Font(name="Calibri", size=11, bold=False)
    font_total = Font(name="Calibri", size=11, bold=True)

    fill_none = PatternFill(fill_type=None)
    fill_total = PatternFill(
        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
    )

    border_thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    align_center = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    COLS_MONEDA = {"VALOR", "SALDO", "CARGOS", "ABONOS", "NETO"}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.views.sheetView[0].showGridLines = True

        if ws.max_row < 1 or ws.max_column < 1:
            continue

        for cell in ws[1]:
            cell.fill = fill_none
            cell.font = font_header
            cell.alignment = align_center
            cell.border = border_thin

        for row_idx in range(2, ws.max_row + 1):
            primera_celda = str(ws.cell(row=row_idx, column=1).value or "").upper()
            es_fila_total = "TOTAL" in primera_celda

            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                header_val = str(ws.cell(row=1, column=col_idx).value or "").upper()

                cell.border = border_thin
                cell.font = font_total if es_fila_total else font_body
                cell.fill = fill_total if es_fila_total else fill_none

                if header_val in COLS_MONEDA or (
                    sheet_name == "Resumen" and col_idx == 2
                ):
                    cell.number_format = FORMATO_MONEDA_IMAGEN
                    cell.alignment = align_right
                elif header_val in ["FECHA", "DCTO", "SUCURSAL", "OBSERVACION"]:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if cell.number_format == FORMATO_MONEDA_IMAGEN and isinstance(
                    cell.value, (int, float)
                ):
                    val_str = f"${cell.value:,.2f}"
                max_len = max(max_len, len(val_str))

            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    wb.save(excel_path)
    wb.close()


class ClasificadorConceptos(tk.Toplevel):

    def __init__(self, parent, excel_path, callback_actualizar_gui):
        super().__init__(parent)
        self.title("Clasificar Conceptos Bancarios")
        self.geometry("600x500")
        self.transient(parent)
        self.grab_set()

        self.excel_path = excel_path
        self.callback_actualizar_gui = callback_actualizar_gui
        self.conceptos_unicos = []
        self.conceptos_bancarios_previos = set()

        self._configurar_estilos()
        self._construir_ui()
        self._cargar_datos()

    def _configurar_estilos(self):
        style = ttk.Style(self)
        style.configure(
            "CheckTree.Treeview", rowheight=30, font=("Segoe UI", 10)
        )
        style.configure(
            "CheckTree.Treeview.Heading", font=("Segoe UI", 10, "bold")
        )

    def _construir_ui(self):
        frame_top = ttk.Frame(self, padding=10)
        frame_top.pack(fill=tk.X)

        ttk.Label(
            frame_top, text="🔍 Buscar concepto:", font=("Segoe UI", 10)
        ).pack(side=tk.LEFT, padx=(0, 10))

        self.var_busqueda = tk.StringVar()
        self.var_busqueda.trace_add("write", self._filtrar_lista)
        entry_buscar = ttk.Entry(
            frame_top, textvariable=self.var_busqueda, font=("Segoe UI", 10)
        )
        entry_buscar.pack(side=tk.LEFT, fill=tk.X, expand=True)

        frame_lista = ttk.Frame(self, padding=10)
        frame_lista.pack(fill=tk.BOTH, expand=True)

        self.tree = ttk.Treeview(
            frame_lista,
            columns=("check", "DESCRIPCIÓN"),
            show="headings",
            style="CheckTree.Treeview",
        )
        self.tree.heading("check", text="Sel")
        self.tree.heading("DESCRIPCIÓN", text="Descripción")

        self.tree.column("check", width=40, anchor="center", stretch=False)
        self.tree.column("DESCRIPCIÓN", anchor="w", stretch=True)

        scrollbar = ttk.Scrollbar(
            frame_lista, orient="vertical", command=self.tree.yview
        )
        self.tree.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.tree.bind("<ButtonRelease-1>", self._toggle_check)

        frame_bottom = ttk.Frame(self, padding=10)
        frame_bottom.pack(fill=tk.X)

        btn_guardar = ttk.Button(
            frame_bottom,
            text="Guardar y Aplicar",
            command=self._aplicar_y_guardar,
        )
        btn_guardar.pack(side=tk.RIGHT, padx=5)

        btn_cancelar = ttk.Button(
            frame_bottom, text="Omitir", command=self._cancelar
        )
        btn_cancelar.pack(side=tk.RIGHT, padx=5)

    def _cargar_datos(self):
        try:
            df = pd.read_excel(self.excel_path, sheet_name="Conceptos")
            if "DESCRIPCIÓN" in df.columns:
                if "Observacion" in df.columns:
                    bancarios_df = df[df["Observacion"].astype(str).str.strip() == "Bancario"]
                    self.conceptos_bancarios_previos = set(bancarios_df["DESCRIPCIÓN"].dropna().astype(str))

                self.conceptos_unicos = (
                    df["DESCRIPCIÓN"].dropna().unique().tolist()
                )
                self._actualizar_treeview(self.conceptos_unicos)
            else:
                messagebox.showwarning(
                    "Aviso",
                    "No se encontró la columna 'DESCRIPCIÓN' en la hoja Conceptos.",
                )
        except Exception as e:
            messagebox.showerror(
                "Error", f"Error al leer el archivo Excel:\n{e}"
            )

    def _actualizar_treeview(self, lista_conceptos):
        for item in self.tree.get_children():
            self.tree.delete(item)

        for concepto in lista_conceptos:
            estado = "✅" if str(concepto) in self.conceptos_bancarios_previos else "🔲"
            self.tree.insert("", tk.END, values=(estado, str(concepto)))

    def _filtrar_lista(self, *args):
        texto_busqueda = self.var_busqueda.get().lower()
        if not texto_busqueda:
            self._actualizar_treeview(self.conceptos_unicos)
            return

        filtrados = [
            c
            for c in self.conceptos_unicos
            if texto_busqueda in str(c).lower()
        ]

        estado_actual = {
            self.tree.item(item, "values")[1]: self.tree.item(item, "values")[0]
            for item in self.tree.get_children()
        }

        self._actualizar_treeview(filtrados)

        for item in self.tree.get_children():
            val = self.tree.item(item, "values")
            if val[1] in estado_actual and estado_actual[val[1]] == "✅":
                self.tree.item(item, values=("✅", val[1]))

    def _toggle_check(self, event):
        item = self.tree.identify_row(event.y)
        column = self.tree.identify_column(event.x)

        if item and column == "#1":
            valores = self.tree.item(item, "values")
            nuevo_estado = "✅" if valores[0] == "🔲" else "🔲"
            self.tree.item(item, values=(nuevo_estado, valores[1]))

    def _aplicar_y_guardar(self):
        conceptos_seleccionados = set()
        for item in self.tree.get_children():
            valores = self.tree.item(item, "values")
            if valores[0] == "✅":
                conceptos_seleccionados.add(valores[1])

        try:
            excel_file = pd.ExcelFile(self.excel_path)
            hojas_dict = {}

            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(excel_file, sheet_name=sheet_name)

                if "DESCRIPCIÓN" in df.columns:
                    if "Observacion" not in df.columns:
                        df["Observacion"] = None

                    mask = (
                        df["DESCRIPCIÓN"]
                        .astype(str)
                        .isin(conceptos_seleccionados)
                    )
                    df.loc[mask, "Observacion"] = "Bancario"
                    
                    unmask = (
                        df["DESCRIPCIÓN"]
                        .astype(str)
                        .isin(self.conceptos_bancarios_previos - conceptos_seleccionados)
                    )
                    df.loc[unmask, "Observacion"] = None

                    # Separar fila de TOTAL GENERAL para dejarla siempre al final
                    es_total = (
                        df["DESCRIPCIÓN"].astype(str).str.upper()
                        == "TOTAL GENERAL"
                    )
                    df_datos = df[~es_total].copy()
                    df_total = df[es_total].copy()

# --- NUEVA LÓGICA DE ORDENAMIENTO ---
                    # 1. Identificar si es bancario (False va primero porque queremos Bancarios arriba)
                    df_datos["_ord_bancario"] = df_datos["Observacion"] != "Bancario"

                    # 2. Identificar negativos arriba y positivos abajo (buscamos columnas numéricas comunes)
                    col_ref_numerica = None
                    for col in ["VALOR", "NETO", "CARGOS"]:
                        if col in df_datos.columns:
                            col_ref_numerica = col
                            break
                    
                    if col_ref_numerica:
                        # Corregido: errors="coerce" para transformar valores no numéricos en NaN y rellenar con 0
                        df_datos[col_ref_numerica] = pd.to_numeric(df_datos[col_ref_numerica], errors="coerce").fillna(0)
                        # True (1) para positivos o ceros, False (0) para negativos -> Negativos quedan arriba
                        df_datos["_ord_signo"] = df_datos[col_ref_numerica] >= 0
                    else:
                        df_datos["_ord_signo"] = False

                    # 3. Orden alfabético A-Z para la descripción
                    df_datos["_ord_desc"] = df_datos["DESCRIPCIÓN"].astype(str).str.upper()

                    # Aplicar orden jerárquico multinivel
                    criterios_orden = ["_ord_bancario", "_ord_signo", "_ord_desc"]
                    columnas_auxiliares = criterios_orden.copy()
                    
                    df_datos = df_datos.sort_values(
                        by=criterios_orden, kind="stable"
                    ).drop(columns=columnas_auxiliares)

                    # Concatenar datos ordenados + Total
                    df = pd.concat([df_datos, df_total], ignore_index=True)

                hojas_dict[sheet_name] = df

            with pd.ExcelWriter(self.excel_path, engine="openpyxl") as writer:
                for sheet_name, df_sheet in hojas_dict.items():
                    df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)

            aplicar_estilos_segun_imagen(self.excel_path)

            messagebox.showinfo(
                "Éxito", "Conceptos ordenados por prioridad bancaria, signo y alfabéticamente."
            )
            self.destroy()
            self.callback_actualizar_gui(self.excel_path)

        except Exception as e:
            messagebox.showerror(
                "Error", f"No se pudo procesar el archivo Excel:\n{e}"
            )

    def _cancelar(self):
        self.destroy()
        self.callback_actualizar_gui(self.excel_path)