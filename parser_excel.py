import re
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def es_formato_fecha(texto):
    """Verifica si un string tiene estructura de fecha corta (ej: '1/02', '01/02', '1-02')."""
    if not texto:
        return False
    texto_str = str(texto).strip()
    return bool(re.search(r"^\d{1,2}[/-]\d{1,2}$", texto_str))


def convertir_a_float_seguro(val):
    """Convierte el valor respetando los decimales reales sin correr la coma/punto."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)

    val_str = str(val).strip().replace("$", "").replace(" ", "")
    if not val_str:
        return 0.0

    # Detección de formato latino vs anglosajón
    if "," in val_str and "." in val_str:
        if val_str.rfind(",") > val_str.rfind("."):
            val_str = val_str.replace(".", "").replace(",", ".")
        else:
            val_str = val_str.replace(",", "")
    elif "," in val_str:
        val_str = val_str.replace(",", ".")

    try:
        return float(val_str)
    except ValueError:
        return 0.0


def extraer_saldos_origen_horizontal(data):
    """Localiza los encabezados de resumen y extrae el valor ubicado en la celda de abajo."""
    saldo_anterior = 0.0
    saldo_actual = 0.0

    for r_idx, fila in enumerate(data):
        for c_idx, celda in enumerate(fila):
            if celda is None:
                continue

            texto = str(celda).strip().upper()

            if "SALDO ANTERIOR" in texto:
                if r_idx + 1 < len(data):
                    val = data[r_idx + 1][c_idx]
                    saldo_anterior = convertir_a_float_seguro(val)

            if "SALDO ACTUAL" in texto:
                if r_idx + 1 < len(data):
                    val = data[r_idx + 1][c_idx]
                    saldo_actual = convertir_a_float_seguro(val)

    return saldo_anterior, saldo_actual


def aplicar_estilos_openpyxl(excel_path):
    """Aplica formato Calibri 11, encabezados sin color de fondo,

    y restablece los colores de texto para valores negativos (rojo) y positivos.
    """
    wb = openpyxl.load_workbook(excel_path)

    # Formato de Moneda Estándar: Positivos normales, Negativos en ROJO entre paréntesis, Cero como guion
    FORMATO_MONEDA = '"$"#,##0.00;[Red]("$"#,##0.00);"-"'

    font_header = Font(name="Calibri", size=11, bold=True)
    font_body = Font(name="Calibri", size=11, bold=False)
    font_total = Font(name="Calibri", size=11, bold=True)

    fill_none = PatternFill(fill_type=None)
    fill_total = PatternFill(
        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
    )

    border_thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    align_center = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    COLS_MONEDA = {
        "VALOR",
        "SALDO",
        "CARGOS",
        "ABONOS",
        "NETO",
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.views.sheetView[0].showGridLines = True

        if ws.max_row < 1 or ws.max_column < 1:
            continue

        # 1. Encabezados (Sin color de fondo, Calibri 11 Negrita)
        for cell in ws[1]:
            cell.fill = fill_none
            cell.font = font_header
            cell.alignment = align_center
            cell.border = border_thin

        # 2. Filas de datos
        for row_idx in range(2, ws.max_row + 1):
            primera_celda = str(ws.cell(row=row_idx, column=1).value or "").upper()
            es_fila_total = "TOTAL" in primera_celda

            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                header_val = str(ws.cell(row=1, column=col_idx).value or "").upper()

                cell.border = border_thin
                cell.font = font_total if es_fila_total else font_body
                cell.fill = fill_total if es_fila_total else fill_none

                # Asignación de formato numérico con color rojo para negativos
                if header_val in COLS_MONEDA or (
                    sheet_name == "Resumen" and col_idx == 2
                ):
                    cell.number_format = FORMATO_MONEDA
                    cell.alignment = align_right
                elif header_val in ["FECHA", "DCTO", "SUCURSAL"]:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left

        # 3. Ajuste automático de ancho de columnas
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if cell.number_format == FORMATO_MONEDA and isinstance(
                    cell.value, (int, float)
                ):
                    val_str = f"${cell.value:,.2f}"
                max_len = max(max_len, len(val_str))

            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    wb.save(excel_path)
    wb.close()


def extraer_datos_desde_excel(excel_input_path, excel_output_path):
    wb_in = openpyxl.load_workbook(excel_input_path, data_only=True)
    ws = wb_in.active

    data = list(ws.iter_rows(values_only=True))
    wb_in.close()

    if not data:
        raise ValueError("El archivo Excel está vacío.")

    # 1. Extraer Saldo Anterior y Saldo Actual correctamente
    saldo_anterior, saldo_actual = extraer_saldos_origen_horizontal(data)

    # 2. Filtrar Movimientos
    movimientos_limpios = []
    en_seccion_movimientos = False

    for fila in data:
        if not fila or not any(
            c is not None and str(c).strip() != "" for c in fila
        ):
            continue

        col_a = str(fila[0]).strip() if fila[0] is not None else ""
        linea_completa = " ".join(
            [str(c).upper() for c in fila if c is not None]
        )

        if "FECHA" in linea_completa and "DESCRIPCI" in linea_completa:
            en_seccion_movimientos = True
            continue

        if not en_seccion_movimientos or not es_formato_fecha(col_a):
            continue

        fecha = col_a
        descripcion = str(fila[1]).strip() if len(fila) > 1 and fila[1] else ""
        sucursal = str(fila[2]).strip() if len(fila) > 2 and fila[2] else ""
        dcto = str(fila[3]).strip() if len(fila) > 3 and fila[3] else ""

        valor = convertir_a_float_seguro(fila[4]) if len(fila) > 4 else 0.0
        saldo = convertir_a_float_seguro(fila[5]) if len(fila) > 5 else 0.0

        movimientos_limpios.append(
            {
                "FECHA": fecha,
                "DESCRIPCIÓN": descripcion,
                "SUCURSAL": sucursal,
                "DCTO": dcto,
                "VALOR": valor,
                "SALDO": saldo,
            }
        )

    df_movimientos = pd.DataFrame(movimientos_limpios)

    total_cargos_conceptos = 0.0
    total_abonos_conceptos = 0.0

    if not df_movimientos.empty:
        # Ordenar: Negativos (Cargos) Primero
        df_movimientos["_es_positivo"] = df_movimientos["VALOR"] >= 0
        df_movimientos = (
            df_movimientos.sort_values(by="_es_positivo", kind="stable")
            .drop(columns=["_es_positivo"])
            .reset_index(drop=True)
        )

        # 3. Hoja Conceptos
        df_calc = df_movimientos.copy()
        df_calc["_CARGOS_TEMP"] = df_calc["VALOR"].apply(
            lambda x: x if x < 0 else 0.0
        )
        df_calc["_ABONOS_TEMP"] = df_calc["VALOR"].apply(
            lambda x: x if x > 0 else 0.0
        )

        df_conceptos = (
            df_calc.groupby("DESCRIPCIÓN", as_index=False)
            .agg(
                CARGOS=("_CARGOS_TEMP", "sum"),
                ABONOS=("_ABONOS_TEMP", "sum"),
                NETO=("VALOR", "sum"),
            )
            .sort_values(by="CARGOS", ascending=True)
        )

        total_cargos_conceptos = float(df_conceptos["CARGOS"].sum())
        total_abonos_conceptos = float(df_conceptos["ABONOS"].sum())

        fila_total = pd.DataFrame(
            [
                {
                    "DESCRIPCIÓN": "TOTAL GENERAL",
                    "CARGOS": total_cargos_conceptos,
                    "ABONOS": total_abonos_conceptos,
                    "NETO": float(df_conceptos["NETO"].sum()),
                }
            ]
        )

        df_conceptos = pd.concat([df_conceptos, fila_total], ignore_index=True)
    else:
        df_conceptos = pd.DataFrame(
            columns=["DESCRIPCIÓN", "CARGOS", "ABONOS", "NETO"]
        )

    # 4. Hoja Resumen en vertical
    datos_resumen = [
        {"CONCEPTO": "SALDO ANTERIOR", "VALOR": saldo_anterior},
        {"CONCEPTO": "CARGOS TOTALES", "VALOR": total_cargos_conceptos},
        {"CONCEPTO": "TOTAL ABONOS", "VALOR": total_abonos_conceptos},
        {"CONCEPTO": "SALDO ACTUAL", "VALOR": saldo_actual},
    ]

    df_resumen = pd.DataFrame(datos_resumen)

    # 5. Exportar a Excel y aplicar estilos
    with pd.ExcelWriter(excel_output_path, engine="openpyxl") as writer:
        df_movimientos.to_excel(writer, sheet_name="Datos", index=False)
        df_resumen.to_excel(writer, sheet_name="Resumen", index=False)
        df_conceptos.to_excel(writer, sheet_name="Conceptos", index=False)

    aplicar_estilos_openpyxl(excel_output_path)

    return excel_output_path